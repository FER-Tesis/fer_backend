import csv
from collections import defaultdict
from datetime import datetime, time, timedelta, timezone
from io import BytesIO, StringIO

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.core.config import settings
from app.repositories import export_job_repository, emotion_event_repository
from app.services.minio_service import MinioService
from app.services.monitoring_service import _fetch_supervisor_agents


PERU_TZ = timezone(timedelta(hours=-5))
UTC_TZ = timezone.utc

EMOTION_ORDER = [
    "neutral",
    "happy",
    "sad",
    "surprise",
    "fear",
    "disgust",
    "anger",
]

EMOTION_LABELS = {
    "neutral": "Neutral",
    "happy": "Felicidad",
    "sad": "Tristeza",
    "surprise": "Sorpresa",
    "fear": "Miedo",
    "disgust": "Disgusto",
    "anger": "Enojo",
}

EMOTION_COLORS = {
    "neutral": "B7BDC6",
    "happy": "F6C85F",
    "sad": "6FA8DC",
    "surprise": "A78BFA",
    "fear": "8E7CC3",
    "disgust": "93C47D",
    "anger": "E06666",
}

GROUP_BY_LABELS = {
    "day": "Diaria",
    "week": "Semanal",
    "month": "Mensual",
}

class ExportDomainError(ValueError):
    pass


def _local_date_range_to_utc(start_date, end_date):
    start_local = datetime.combine(start_date, time.min).replace(tzinfo=PERU_TZ)
    end_local = datetime.combine(end_date, time.max).replace(tzinfo=PERU_TZ)

    return start_local.astimezone(UTC_TZ), end_local.astimezone(UTC_TZ)


def _to_peru(dt: datetime) -> datetime:
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=UTC_TZ)
    return dt.astimezone(PERU_TZ)


def _build_period_key(dt: datetime, group_by: str) -> str:
    dt_peru = _to_peru(dt)

    if group_by == "day":
        return dt_peru.strftime("%Y-%m-%d")

    if group_by == "week":
        year, week, _ = dt_peru.isocalendar()
        return f"{year}-W{week:02d}"

    if group_by == "month":
        return dt_peru.strftime("%Y-%m")

    raise ExportDomainError("invalid_group_by")


def _build_csv_bytes(rows: list[dict]) -> bytes:
    if not rows:
        return b"Sin datos\n"

    buffer = StringIO()
    writer = csv.DictWriter(buffer, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    writer.writerows(rows)
    return buffer.getvalue().encode("utf-8-sig")

def _get_record_agent_id(record: dict) -> str:
    return str(record.get("agent_id") or record.get("agentId") or "")


def _get_agent_name(agent: dict) -> str:
    return (
        agent.get("name")
        or agent.get("full_name")
        or agent.get("email")
        or str(agent.get("id"))
    )


def _build_agent_summary_rows(records: list[dict], agents: list[dict]) -> list[dict]:
    agent_names = {str(agent["id"]): _get_agent_name(agent) for agent in agents}

    buckets = defaultdict(
        lambda: {
            "agent_name": "",
            "total_records": 0,
            "emotion_counts": {emotion: 0 for emotion in EMOTION_ORDER},
        }
    )

    for record in records:
        agent_id = _get_record_agent_id(record)
        emotion = record.get("emotion")

        buckets[agent_id]["agent_name"] = agent_names.get(agent_id, agent_id)
        buckets[agent_id]["total_records"] += 1

        if emotion in EMOTION_ORDER:
            buckets[agent_id]["emotion_counts"][emotion] += 1

    rows = []

    for agent_id, bucket in buckets.items():
        total = bucket["total_records"]
        emotion_counts = bucket["emotion_counts"]

        dominant_emotion = max(
            EMOTION_ORDER,
            key=lambda emotion: emotion_counts[emotion],
        )

        row = {
            "agent_id": agent_id,
            "agent_name": bucket["agent_name"],
            "total_records": total,
            "dominant_emotion": EMOTION_LABELS.get(dominant_emotion, dominant_emotion),
        }

        for emotion in EMOTION_ORDER:
            row[f"{emotion}_count"] = emotion_counts[emotion]
            row[f"{emotion}_percentage"] = round(
                emotion_counts[emotion] / total * 100, 2
            ) if total else 0

        rows.append(row)

    return sorted(rows, key=lambda row: row["total_records"], reverse=True)


def _style_title(ws, title: str, subtitle: str = ""):
    ws.merge_cells("A1:H1")
    ws["A1"] = title
    ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F2937")
    ws["A1"].alignment = Alignment(horizontal="center")

    if subtitle:
        ws.merge_cells("A2:H2")
        ws["A2"] = subtitle
        ws["A2"].font = Font(size=11, color="374151")
        ws["A2"].alignment = Alignment(horizontal="center")


def _style_header(row):
    for cell in row:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="374151")
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=Side(style="thin", color="D1D5DB"))


def _autofit(ws):
    for column_cells in ws.columns:
        max_length = 0
        column = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = cell.value
            if value is not None:
                max_length = max(max_length, len(str(value)))

        ws.column_dimensions[column].width = min(max_length + 3, 35)


def _append_table(ws, start_row: int, headers: list[str], rows: list[dict]):
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=start_row, column=col_idx, value=header)

    _style_header(ws[start_row])

    for row_idx, row in enumerate(rows, start_row + 1):
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(header))

    return start_row + len(rows)

def _get_emotion_key_from_label(label: str) -> str | None:
    for key, value in EMOTION_LABELS.items():
        if value == label:
            return key

    return None

def _build_agent_detail_sheet(wb: Workbook, detail_rows: list[dict]) -> None:
    ws = wb.create_sheet("Detalle por agente")

    _style_title(
        ws,
        "Detalle emocional por agente",
        "Resumen comparativo de emociones detectadas por integrante del equipo",
    )

    current_row = 4

    # =========================
    # BLOQUE 1: RESUMEN GENERAL
    # =========================
    ws.cell(row=current_row, column=1, value="Resumen general")
    ws.cell(row=current_row, column=1).font = Font(size=14, bold=True, color="111827")
    current_row += 2

    summary_headers = [
        "Agente",
        "Total registros",
        "Emoción predominante",
    ]

    summary_rows = [
        {
            "Agente": row["agent_name"],
            "Total registros": row["total_records"],
            "Emoción predominante": row["dominant_emotion"],
        }
        for row in detail_rows
    ]

    _append_table(ws, current_row, summary_headers, summary_rows)

    dominant_col = summary_headers.index("Emoción predominante") + 1

    for row_idx in range(current_row + 1, current_row + 1 + len(summary_rows)):
        cell = ws.cell(row=row_idx, column=dominant_col)
        emotion_key = _get_emotion_key_from_label(cell.value)

        if emotion_key:
            cell.fill = PatternFill(
                "solid",
                fgColor=EMOTION_COLORS.get(emotion_key, "FFFFFF"),
            )

        cell.font = Font(bold=True, color="111827")
        cell.alignment = Alignment(horizontal="center")

    current_row += len(summary_rows) + 4

    # ==============================
    # BLOQUE 2: CANTIDAD DE EMOCIONES
    # ==============================
    ws.cell(row=current_row, column=1, value="Cantidad de emociones")
    ws.cell(row=current_row, column=1).font = Font(size=14, bold=True, color="111827")
    current_row += 2

    count_headers = ["Agente"] + [EMOTION_LABELS[emotion] for emotion in EMOTION_ORDER]

    count_rows = []

    for row in detail_rows:
        count_row = {
            "Agente": row["agent_name"],
        }

        for emotion in EMOTION_ORDER:
            count_row[EMOTION_LABELS[emotion]] = row.get(f"{emotion}_count", 0)

        count_rows.append(count_row)

    _append_table(ws, current_row, count_headers, count_rows)

    for col_idx, emotion in enumerate(EMOTION_ORDER, start=2):
        ws.cell(row=current_row, column=col_idx).fill = PatternFill(
            "solid",
            fgColor=EMOTION_COLORS[emotion],
        )
        ws.cell(row=current_row, column=col_idx).font = Font(bold=True, color="111827")

    current_row += len(count_rows) + 4

    # ==========================
    # BLOQUE 3: PORCENTAJES
    # ==========================
    ws.cell(row=current_row, column=1, value="Porcentaje de emociones")
    ws.cell(row=current_row, column=1).font = Font(size=14, bold=True, color="111827")
    current_row += 2

    percentage_headers = ["Agente"] + [
        f"{EMOTION_LABELS[emotion]} %"
        for emotion in EMOTION_ORDER
    ]

    percentage_rows = []

    for row in detail_rows:
        percentage_row = {
            "Agente": row["agent_name"],
        }

        for emotion in EMOTION_ORDER:
            percentage_row[f"{EMOTION_LABELS[emotion]} %"] = row.get(
                f"{emotion}_percentage",
                0,
            )

        percentage_rows.append(percentage_row)

    _append_table(ws, current_row, percentage_headers, percentage_rows)

    for row_idx in range(current_row + 1, current_row + 1 + len(percentage_rows)):
        for col_idx in range(2, len(percentage_headers) + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = '0.00"%"'

    for col_idx in range(2, len(percentage_headers) + 1):
        ws.cell(row=current_row, column=col_idx).fill = PatternFill(
            "solid",
            fgColor="DBEAFE",
        )
        ws.cell(row=current_row, column=col_idx).font = Font(bold=True, color="111827")

    ws.freeze_panes = "A5"
    _autofit(ws)

LOW_SAMPLE_THRESHOLD = 30


def _format_percentage(value: float) -> str:
    return f"{value:.2f}%"


def _get_dominant_emotion_from_row(row: dict) -> tuple[str, int, float]:
    total = row.get("total_records", 0)

    dominant = max(
        EMOTION_ORDER,
        key=lambda emotion: row.get(f"{emotion}_count", 0),
    )

    count = row.get(f"{dominant}_count", 0)
    percentage = round((count / total * 100), 2) if total else 0

    return dominant, count, percentage


def _build_period_evolution_sheet(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("Evolución del período")

    _style_title(
        ws,
        "Evolución del período",
        "Resumen del comportamiento emocional por período analizado",
    )

    current_row = 4

    ws.cell(row=current_row, column=1, value="Resumen por período")
    ws.cell(row=current_row, column=1).font = Font(size=14, bold=True, color="111827")

    current_row += 2

    summary_headers = [
        "Período",
        "Total registros",
        "Emoción predominante",
        "% predominante",
        "Observación",
    ]

    summary_rows = []

    for row in rows:
        dominant_emotion, _, dominant_percentage = _get_dominant_emotion_from_row(row)
        total_records = row.get("total_records", 0)

        summary_rows.append({
            "Período": row.get("period", "-"),
            "Total registros": total_records,
            "Emoción predominante": EMOTION_LABELS.get(dominant_emotion, dominant_emotion),
            "% predominante": dominant_percentage,
            "Observación": (
                "Muestra reducida"
                if total_records < LOW_SAMPLE_THRESHOLD
                else "Muestra suficiente"
            ),
        })

    _append_table(ws, current_row, summary_headers, summary_rows)

    dominant_col = summary_headers.index("Emoción predominante") + 1
    percentage_col = summary_headers.index("% predominante") + 1
    observation_col = summary_headers.index("Observación") + 1

    for row_idx in range(current_row + 1, current_row + 1 + len(summary_rows)):
        dominant_cell = ws.cell(row=row_idx, column=dominant_col)
        emotion_key = _get_emotion_key_from_label(dominant_cell.value)

        if emotion_key:
            dominant_cell.fill = PatternFill(
                "solid",
                fgColor=EMOTION_COLORS.get(emotion_key, "FFFFFF"),
            )

        dominant_cell.font = Font(bold=True, color="111827")
        dominant_cell.alignment = Alignment(horizontal="center")

        ws.cell(row=row_idx, column=percentage_col).number_format = '0.00"%"'

        observation_cell = ws.cell(row=row_idx, column=observation_col)

        if observation_cell.value == "Muestra reducida":
            observation_cell.fill = PatternFill("solid", fgColor="FEF3C7")
            observation_cell.font = Font(bold=True, color="92400E")
        else:
            observation_cell.fill = PatternFill("solid", fgColor="DCFCE7")
            observation_cell.font = Font(bold=True, color="166534")

    current_row += len(summary_rows) + 4

    ws.cell(row=current_row, column=1, value="Datos para gráfico")
    ws.cell(row=current_row, column=1).font = Font(size=14, bold=True, color="111827")

    current_row += 2

    chart_headers = ["Período"] + [EMOTION_LABELS[emotion] for emotion in EMOTION_ORDER]

    chart_rows = []

    for row in rows:
        chart_row = {
            "Período": row.get("period", "-"),
        }

        for emotion in EMOTION_ORDER:
            chart_row[EMOTION_LABELS[emotion]] = row.get(f"{emotion}_percentage", 0)

        chart_rows.append(chart_row)

    chart_table_start = current_row

    _append_table(ws, chart_table_start, chart_headers, chart_rows)

    for row_idx in range(chart_table_start + 1, chart_table_start + 1 + len(chart_rows)):
        for col_idx in range(2, len(chart_headers) + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = '0.00"%"'

    for col_idx, emotion in enumerate(EMOTION_ORDER, start=2):
        ws.cell(row=chart_table_start, column=col_idx).fill = PatternFill(
            "solid",
            fgColor=EMOTION_COLORS[emotion],
        )
        ws.cell(row=chart_table_start, column=col_idx).font = Font(
            bold=True,
            color="111827",
        )

    if rows:
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "percentStacked"
        chart.overlap = 100
        chart.title = "Distribución emocional por período"
        chart.style = 10
        chart.width = 18
        chart.height = 9
        chart.legend.position = "r"

        data = Reference(
            ws,
            min_col=2,
            max_col=len(chart_headers),
            min_row=chart_table_start,
            max_row=chart_table_start + len(chart_rows),
        )

        cats = Reference(
            ws,
            min_col=1,
            min_row=chart_table_start + 1,
            max_row=chart_table_start + len(chart_rows),
        )

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        for idx, emotion in enumerate(EMOTION_ORDER):
            if idx < len(chart.series):
                chart.series[idx].graphicalProperties.solidFill = EMOTION_COLORS[emotion]
                chart.series[idx].graphicalProperties.line.solidFill = EMOTION_COLORS[emotion]

        ws.add_chart(chart, "I5")

    note_row = chart_table_start + len(chart_rows) + 3

    ws.merge_cells(
        start_row=note_row,
        start_column=1,
        end_row=note_row,
        end_column=8,
    )

    note_cell = ws.cell(row=note_row, column=1)
    note_cell.value = (
        f"Nota: los períodos con menos de {LOW_SAMPLE_THRESHOLD} registros "
        "deben interpretarse con cautela por el tamaño reducido de la muestra."
    )
    note_cell.fill = PatternFill("solid", fgColor="FEF3C7")
    note_cell.font = Font(italic=True, color="92400E")
    note_cell.alignment = Alignment(wrap_text=True)

    ws.freeze_panes = "A6"
    _autofit(ws)


def _build_xlsx_bytes(
    rows: list[dict],
    *,
    export_type: str,
    params: dict,
    detail_rows: list[dict] | None = None,
) -> bytes:
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Resumen"

    title = "Reporte emocional"
    if export_type == "team":
        title = "Reporte emocional del equipo"
    elif export_type == "agent":
        title = "Reporte emocional del agente"
    elif export_type == "current":
        title = "Reporte de estado actual"

    subtitle = f"Generado el {_to_peru(datetime.now(timezone.utc)).strftime('%Y-%m-%d %H:%M')}"
    _style_title(ws_summary, title, subtitle)

    total_records = sum(row.get("total_records", 0) for row in rows)

    emotion_totals = {emotion: 0 for emotion in EMOTION_ORDER}
    for row in rows:
        for emotion in EMOTION_ORDER:
            emotion_totals[emotion] += row.get(f"{emotion}_count", 0)

    dominant_emotion = (
        max(EMOTION_ORDER, key=lambda emotion: emotion_totals[emotion])
        if total_records
        else None
    )

    kpis = [
        ("Total de registros", total_records),
        (
            "Emoción dominante",
            EMOTION_LABELS.get(dominant_emotion, "Sin datos") if dominant_emotion else "Sin datos",
        ),
        ("Agrupación", GROUP_BY_LABELS.get(params.get("group_by"), "-")),
        ("Periodo", f"{params.get('start_date', '-')} a {params.get('end_date', '-')}"),
    ]

    if export_type == "team":
        kpis.append(("Agentes analizados", len(detail_rows) if detail_rows else 0))

    ws_summary["A4"] = "Indicadores principales"
    ws_summary["A4"].font = Font(size=14, bold=True, color="111827")

    row_cursor = 6
    for idx, (label, value) in enumerate(kpis):
        col = 1 + (idx % 2) * 4
        row = row_cursor + (idx // 2) * 3

        ws_summary.cell(row=row, column=col, value=label)
        ws_summary.cell(row=row + 1, column=col, value=value)

        ws_summary.cell(row=row, column=col).font = Font(bold=True, color="4B5563")
        ws_summary.cell(row=row + 1, column=col).font = Font(size=14, bold=True, color="111827")
        ws_summary.cell(row=row + 1, column=col).fill = PatternFill("solid", fgColor="E5E7EB")
        ws_summary.cell(row=row + 1, column=col).alignment = Alignment(horizontal="center")

    ws_summary["A15"] = "Distribución general de emociones"
    ws_summary["A15"].font = Font(size=14, bold=True)

    emotion_table_start = 17

    emotion_table_rows = []
    for emotion in EMOTION_ORDER:
        count = emotion_totals[emotion]
        percentage = round(count / total_records * 100, 2) if total_records else 0

        emotion_table_rows.append({
            "Emoción": EMOTION_LABELS.get(emotion, emotion),
            "Cantidad": count,
            "Porcentaje": percentage,
        })

    _append_table(
        ws_summary,
        emotion_table_start,
        ["Emoción", "Cantidad", "Porcentaje"],
        emotion_table_rows,
    )

    for idx, emotion in enumerate(EMOTION_ORDER, start=emotion_table_start + 1):
        ws_summary.cell(row=idx, column=1).fill = PatternFill(
            "solid",
            fgColor=EMOTION_COLORS[emotion],
        )
        ws_summary.cell(row=idx, column=3).number_format = '0.00"%"'

    if total_records:
        chart = BarChart()
        chart.title = "Emociones detectadas"
        chart.style = 10
        chart.legend = None
        chart.width = 14
        chart.height = 8

        data = Reference(
            ws_summary,
            min_col=2,
            min_row=emotion_table_start,
            max_row=emotion_table_start + len(EMOTION_ORDER),
        )

        cats = Reference(
            ws_summary,
            min_col=1,
            min_row=emotion_table_start + 1,
            max_row=emotion_table_start + len(EMOTION_ORDER),
        )

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        ws_summary.add_chart(chart, "E16")

    if rows:
        _build_period_evolution_sheet(wb, rows)

    if detail_rows:
        _build_agent_detail_sheet(wb, detail_rows)

    for ws in wb.worksheets:
        if ws.title != "Detalle por agente":
            ws.freeze_panes = "A2"

        _autofit(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def _aggregate_records(records: list[dict], group_by: str, aggregates: list[str]) -> list[dict]:
    buckets = defaultdict(
        lambda: {
            "total_records": 0,
            "emotion_counts": {emotion: 0 for emotion in EMOTION_ORDER},
        }
    )

    for record in records:
        period = _build_period_key(record["timestamp"], group_by)
        emotion = record["emotion"]

        buckets[period]["total_records"] += 1

        if emotion in buckets[period]["emotion_counts"]:
            buckets[period]["emotion_counts"][emotion] += 1

    rows = []

    for period in sorted(buckets.keys()):
        bucket = buckets[period]
        total_records = bucket["total_records"]
        emotion_counts = bucket["emotion_counts"]

        row = {
            "period": period,
            "total_records": total_records,
        }

        if "emotion_count" in aggregates:
            for emotion in EMOTION_ORDER:
                row[f"{emotion}_count"] = emotion_counts[emotion]

        if "emotion_percentage" in aggregates:
            for emotion in EMOTION_ORDER:
                percentage = (emotion_counts[emotion] / total_records * 100) if total_records else 0
                row[f"{emotion}_percentage"] = round(percentage, 2)

        if "dominant_emotion" in aggregates:
            dominant_emotion = max(
                EMOTION_ORDER,
                key=lambda emotion: emotion_counts[emotion],
            )
            row["dominant_emotion"] = dominant_emotion

        rows.append(row)

    return rows


def _build_current_rows(snapshot: list[dict]) -> list[dict]:
    return [
        {
            "name": row["name"],
            "emotion_label": row["emotion_label"],
            "updated_at": row["updated_at"],
        }
        for row in snapshot
    ]


async def _resolve_supervisor_agent_ids(supervisor_id: str) -> list[str]:
    agents = await _fetch_supervisor_agents(supervisor_id)
    return [str(agent["id"]) for agent in agents]


async def create_export_job(supervisor_id: str, payload) -> str:
    params: dict = {}

    if payload.type == "current":
        params = {
            "snapshot": [item.model_dump() for item in payload.snapshot],
        }

    elif payload.type == "team":
        params = {
            "start_date": payload.start_date.isoformat(),
            "end_date": payload.end_date.isoformat(),
            "group_by": payload.group_by,
            "aggregates": payload.aggregates,
        }

    elif payload.type == "agent":
        params = {
            "agent_id": payload.agent_id,
            "start_date": payload.start_date.isoformat(),
            "end_date": payload.end_date.isoformat(),
            "group_by": payload.group_by,
            "aggregates": payload.aggregates,
        }

    document = {
        "requested_by": supervisor_id,
        "requested_at": datetime.now(timezone.utc),
        "status": "processing",
        "type": payload.type,
        "format": payload.format,
        "params": params,
        "file": None,
        "error": None,
        "completed_at": None,
    }

    return await export_job_repository.create_export_job(document)


async def process_export_job(job_id: str):
    minio_service = MinioService()

    try:
        minio_service.ensure_bucket()

        job = await export_job_repository.get_export_job_by_id(job_id)
        if not job:
            return

        export_type = job["type"]
        export_format = job["format"]
        params = job["params"]
        requested_by = job["requested_by"]

        rows: list[dict] = []
        detail_rows: list[dict] | None = None
        raw_rows: list[dict] | None = None
        file_name = ""

        if export_type == "current":
            rows = _build_current_rows(params["snapshot"])
            file_name = f"current_export_{job_id}.{export_format}"

        elif export_type == "team":
            agents = await _fetch_supervisor_agents(requested_by)
            supervisor_agent_ids = [str(agent["id"]) for agent in agents]

            start_utc, end_utc = _local_date_range_to_utc(
                datetime.fromisoformat(params["start_date"]).date(),
                datetime.fromisoformat(params["end_date"]).date(),
            )

            events = await emotion_event_repository.get_emotion_events_between(
                start=start_utc,
                end=end_utc,
                agent_ids=supervisor_agent_ids,
            )

            rows = _aggregate_records(
                records=events,
                group_by=params["group_by"],
                aggregates=params["aggregates"],
            )

            detail_rows = _build_agent_summary_rows(events, agents)

            raw_rows = [
                {
                    "agent_id": _get_record_agent_id(event),
                    "timestamp": _to_peru(event["timestamp"]).strftime("%Y-%m-%d %H:%M:%S"),
                    "emotion": EMOTION_LABELS.get(event.get("emotion"), event.get("emotion")),
                }
                for event in events
            ]

            file_name = (
                f"team_emotions_{params['start_date']}_{params['end_date']}.{export_format}"
            )

        elif export_type == "agent":
            start_utc, end_utc = _local_date_range_to_utc(
                datetime.fromisoformat(params["start_date"]).date(),
                datetime.fromisoformat(params["end_date"]).date(),
            )

            events = await emotion_event_repository.get_emotion_events_between(
                start=start_utc,
                end=end_utc,
                agent_ids=[params["agent_id"]],
            )

            rows = _aggregate_records(
                records=events,
                group_by=params["group_by"],
                aggregates=params["aggregates"],
            )

            raw_rows = [
                {
                    "agent_id": _get_record_agent_id(event),
                    "timestamp": _to_peru(event["timestamp"]).strftime("%Y-%m-%d %H:%M:%S"),
                    "emotion": EMOTION_LABELS.get(event.get("emotion"), event.get("emotion")),
                }
                for event in events
            ]

            file_name = (
                f"agent_emotions_{params['agent_id']}_{params['start_date']}_{params['end_date']}.{export_format}"
            )

        else:
            raise ExportDomainError("invalid_export_type")

        if export_format == "csv":
            file_bytes = _build_csv_bytes(rows)
            content_type = "text/csv"
        elif export_format == "xlsx":
            file_bytes = _build_xlsx_bytes(
                rows,
                export_type=export_type,
                params=params,
                detail_rows=detail_rows,
            )
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            raise ExportDomainError("invalid_export_format")

        object_key = f"jobs/{job_id}/{file_name}"

        minio_service.upload_bytes(
            object_key=object_key,
            payload=file_bytes,
            content_type=content_type,
        )

        expires_at = datetime.now(timezone.utc) + timedelta(
            minutes=settings.EXPORT_FILE_TTL_MINUTES
        )

        await export_job_repository.mark_export_job_completed(
            job_id=job_id,
            file_data={
                "bucket": settings.MINIO_BUCKET,
                "object_key": object_key,
                "file_name": file_name,
                "content_type": content_type,
                "expires_at": expires_at,
            },
        )

    except Exception as e:
        await export_job_repository.mark_export_job_failed(job_id, str(e))


async def get_export_job(job_id: str):
    return await export_job_repository.get_export_job_by_id(job_id)